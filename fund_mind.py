"""
fund_mind.py

Downloads fund documents from fundinfo.com for each ISIN in funds.xlsx.
Fetches all available document types (PR, MR, KIID, KFS, LM, PRP, AR, etc.)
Language preference: English > German > French > other.
Updates funds.xlsx with columns indicating which documents were downloaded.
"""

import asyncio
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
import openpyxl


# ============================================================
# Logging
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)


# ============================================================
# Configuration
# ============================================================

@dataclass(frozen=True)
class Config:
    excel_file: Path = Path("funds.xlsx")
    output_dir: Path = Path("fund_document_library")

    fundinfo_profile: str = "CH-prof"
    lang_pref: tuple[str, ...] = ("EN", "DE", "FR", "IT", "ES")

    max_concurrency: int = 5
    request_timeout: int = 30
    retry_count: int = 3
    polite_delay_seconds: float = 0.3

    headers: dict[str, str] = field(default_factory=lambda: {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Language": "en-US,en;q=0.9",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": "https://www.fundinfo.com/",
    })

    @property
    def cookies(self) -> dict[str, str]:
        return {
            "DU": self.fundinfo_profile,
            "PrivacyPolicy": "true",
        }

    @property
    def api_url(self) -> str:
        return f"https://www.fundinfo.com/en/{self.fundinfo_profile}/LandingPage/Data"

    @property
    def fund_data_url(self) -> str:
        return f"https://www.fundinfo.com/en/{self.fundinfo_profile}/fund/Data"


# ============================================================
# Data models
# ============================================================

@dataclass(frozen=True)
class Fund:
    isin: str
    name: str = ""


@dataclass(frozen=True)
class Document:
    doc_type: str
    language: str
    date: str
    url: str
    active: bool


@dataclass
class DownloadResult:
    fund: Fund
    downloaded_files: dict[str, str] = field(default_factory=dict)
    available_doc_types: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    legal_form: str = ""
    ter_excl_performance_fee: str = ""
    ept_valuation_frequency: str = ""
    ept_investment_objective: str = ""
    manco: str = ""
    fund_domicile: str = ""
    fund_launch_date: str = ""
    fund_currency: str = ""


# ============================================================
# Constants
# ============================================================

DOC_TYPE_LABELS: dict[str, str] = {
    "PR":   "Prospectus",
    "MR":   "Monthly Report",
    "KFS":  "KFS",
    "KIID": "KIID",
    "KMS":  "KMS",
    "KPP":  "KPP",
    "PRP":  "PRIIPs",
    "LM":   "Legal Message",
    "SUP":  "Supplement",
    "AR":   "Annual Report",
    "ESC":  "ESG Scorecard",
}


# ============================================================
# Utility functions
# ============================================================

def parse_date(value: str) -> datetime:
    if not value:
        return datetime.min
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return datetime.min


def sanitize_filename_part(value: str) -> str:
    value = str(value).strip() if value else ""
    value = re.sub(r'[<>:"/\\|?*]+', "_", value)
    value = re.sub(r"\s+", "_", value)
    return value.strip("._") or "unknown"


def make_output_filename(isin: str, doc_type: str, language: str, doc_date: str) -> str:
    return (
        f"{sanitize_filename_part(isin)}_"
        f"{sanitize_filename_part(doc_type)}_"
        f"{sanitize_filename_part(language or 'XX')}_"
        f"{sanitize_filename_part(doc_date or 'unknown_date')}.pdf"
    )


def pick_best_document(docs: list[Document], lang_pref: tuple[str, ...]) -> Document | None:
    if not docs:
        return None
    active = sorted([d for d in docs if d.active], key=lambda d: parse_date(d.date), reverse=True)
    inactive = sorted([d for d in docs if not d.active], key=lambda d: parse_date(d.date), reverse=True)
    candidates = active or inactive
    for lang in lang_pref:
        for doc in candidates:
            if doc.language.upper() == lang:
                return doc
    return candidates[0] if candidates else None


# ============================================================
# Excel I/O
# ============================================================

def load_funds(path: Path) -> list[Fund]:
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

    if "IDINSTRUMENT" not in headers:
        raise ValueError("Column 'IDINSTRUMENT' not found in Excel file.")

    isin_idx = headers.index("IDINSTRUMENT")
    name_idx = headers.index("DESCRIPTION") if "DESCRIPTION" in headers else None

    funds: list[Fund] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        isin = str(row[isin_idx]).strip() if isin_idx < len(row) and row[isin_idx] else ""
        name = str(row[name_idx]).strip() if name_idx is not None and name_idx < len(row) and row[name_idx] else ""
        if isin:
            funds.append(Fund(isin=isin, name=name))
    return funds


def write_results_to_excel(path: Path, results: list[DownloadResult]) -> None:
    result_map = {r.fund.isin: r for r in results}

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

    if "IDINSTRUMENT" not in headers:
        raise ValueError("Column 'IDINSTRUMENT' not found in Excel file.")

    isin_col_idx = headers.index("IDINSTRUMENT") + 1
    other_doc_types = sorted(k for k in DOC_TYPE_LABELS if k != "PR")
    all_doc_types = ["PR"] + other_doc_types

    metadata_columns = [
        ("Legal Form", "legal_form"),
        ("TER Excluding Performance Fee", "ter_excl_performance_fee"),
        ("EPT Valuation Frequency", "ept_valuation_frequency"),
        ("EPT Investment Objective", "ept_investment_objective"),
        ("ManCo", "manco"),
        ("Fund Domicile Alpha-2", "fund_domicile"),
        ("Fund Launch Date", "fund_launch_date"),
        ("Fund Currency", "fund_currency"),
    ]

    for label, _ in metadata_columns:
        if label not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=label)
            headers.append(label)

    for doc_type in all_doc_types:
        label = DOC_TYPE_LABELS[doc_type]
        if label not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=label)
            headers.append(label)

    for row_idx in range(2, ws.max_row + 1):
        isin_val = ws.cell(row=row_idx, column=isin_col_idx).value
        isin = str(isin_val).strip() if isin_val else ""
        result = result_map.get(isin)
        if not result:
            continue
        for label, attr in metadata_columns:
            ws.cell(row=row_idx, column=headers.index(label) + 1, value=getattr(result, attr, ""))
        for doc_type in all_doc_types:
            label = DOC_TYPE_LABELS[doc_type]
            ws.cell(row=row_idx, column=headers.index(label) + 1, value=result.downloaded_files.get(doc_type, ""))

    wb.save(path)
    logger.info("Excel updated: %s", path)


# ============================================================
# Fundinfo API
# ============================================================

def _match_record_by_isin(records: list[dict[str, Any]], isin: str) -> dict[str, Any]:
    isin_upper = isin.upper()
    for record in records:
        record_isin = str(
            record.get("IDINSTRUMENT") or record.get("ISIN") or record.get("Isin") or ""
        ).strip().upper()
        if record_isin == isin_upper:
            return record
    return records[0]


async def fetch_fund_metadata(
    client: httpx.AsyncClient, isin: str, config: Config
) -> dict[str, str]:
    """Fetch fund-level metadata (legal form, fees, EPT fields) from the fund/Data endpoint."""
    fund_headers = {**config.headers, "Referer": f"https://www.fundinfo.com/en/{config.fundinfo_profile}/fund#OFST020000={isin}"}
    try:
        response = await client.get(
            config.fund_data_url, params={"OFST020000": isin}, headers=fund_headers, timeout=config.request_timeout
        )
        response.raise_for_status()
        payload = response.json()
    except Exception as exc:
        logger.warning("Metadata fetch error for %s: %s", isin, exc)
        return {}

    s = payload.get("Data", {}).get("S", {})
    if not isinstance(s, dict):
        return {}
    return {
        "legal_form": str(s.get("OFST160100") or "").strip(),
        "ter_excl_performance_fee": str(s.get("OFST452100") or "").strip(),
        "ept_valuation_frequency": str(s.get("OFEP010100") or "").strip(),
        "ept_investment_objective": str(s.get("OFEP040400") or "").strip(),
        "manco": str(s.get("OFST001020") or "").strip(),
        "fund_domicile": str(s.get("OFST010010") or "").strip(),
        "fund_launch_date": str(s.get("OFST010240") or "").strip(),
        "fund_currency": str(s.get("OFST010410") or "").strip(),
    }


async def fetch_documents_for_isin(
    client: httpx.AsyncClient, isin: str, config: Config
) -> dict[str, list[Document]]:
    params = {"skip": 0, "query": isin, "orderdirection": ""}
    try:
        response = await client.get(
            config.api_url, params=params, headers=config.headers, timeout=config.request_timeout
        )
        response.raise_for_status()
        payload = response.json()
    except Exception as exc:
        logger.error("API error for %s: %s", isin, exc)
        return {}

    records = payload.get("Data", [])
    if not records:
        logger.warning("No API data returned for %s", isin)
        return {}

    raw_doc_section = _match_record_by_isin(records, isin).get("D") or {}
    if not isinstance(raw_doc_section, dict):
        return {}

    parsed: dict[str, list[Document]] = {}
    for doc_type, raw_docs in raw_doc_section.items():
        if not isinstance(raw_docs, list):
            continue
        docs = [
            Document(
                doc_type=str(doc_type),
                language=str(raw.get("Language") or "").strip(),
                date=str(raw.get("Date") or "").strip(),
                url=str(raw.get("Url") or "").strip(),
                active=bool(raw.get("Active")),
            )
            for raw in raw_docs
            if isinstance(raw, dict)
        ]
        if docs:
            parsed[doc_type] = docs
    return parsed


async def download_pdf(
    client: httpx.AsyncClient, url: str, dest: Path, config: Config
) -> bool:
    for attempt in range(1, config.retry_count + 1):
        try:
            response = await client.get(
                url, headers=config.headers, timeout=config.request_timeout, follow_redirects=True
            )
            if response.status_code == 200 and response.content[:4] == b"%PDF":
                dest.write_bytes(response.content)
                return True
            logger.warning("Bad response for %s | status=%s | attempt=%s", url, response.status_code, attempt)
        except Exception as exc:
            logger.warning("Download error for %s | attempt=%s | error=%s", url, attempt, exc)
        await asyncio.sleep(attempt)
    return False


# ============================================================
# Pipeline
# ============================================================

async def process_fund(
    fund: Fund,
    client: httpx.AsyncClient,
    config: Config,
    semaphore: asyncio.Semaphore,
) -> DownloadResult:
    async with semaphore:
        logger.info("Processing | ISIN=%s | Name=%s", fund.isin, fund.name)
        result = DownloadResult(fund=fund)

        metadata = await fetch_fund_metadata(client, fund.isin, config)
        result.legal_form = metadata.get("legal_form", "")
        result.ter_excl_performance_fee = metadata.get("ter_excl_performance_fee", "")
        result.ept_valuation_frequency = metadata.get("ept_valuation_frequency", "")
        result.ept_investment_objective = metadata.get("ept_investment_objective", "")
        result.manco = metadata.get("manco", "")
        result.fund_domicile = metadata.get("fund_domicile", "")
        result.fund_launch_date = metadata.get("fund_launch_date", "")
        result.fund_currency = metadata.get("fund_currency", "")

        documents_by_type = await fetch_documents_for_isin(client, fund.isin, config)
        if not documents_by_type:
            result.errors.append("No documents found")
            return result

        result.available_doc_types = sorted(documents_by_type.keys())
        fund_dir = config.output_dir / sanitize_filename_part(fund.isin)
        fund_dir.mkdir(parents=True, exist_ok=True)

        for doc_type, docs in documents_by_type.items():
            best = pick_best_document(docs, config.lang_pref)
            if not best or not best.url:
                if best:
                    result.errors.append(f"{doc_type}: missing URL")
                continue

            filename = make_output_filename(fund.isin, best.doc_type, best.language, best.date)
            dest = fund_dir / filename

            if dest.exists():
                logger.info("[%s] Already exists: %s", doc_type, filename)
                result.downloaded_files[doc_type] = filename
                continue

            logger.info("[%s] Downloading | lang=%s | date=%s | file=%s", doc_type, best.language, best.date, filename)
            ok = await download_pdf(client, best.url, dest, config)
            if ok:
                result.downloaded_files[doc_type] = filename
                logger.info("[%s] OK", doc_type)
            else:
                result.errors.append(f"{doc_type}: download failed")
                logger.error("[%s] FAILED", doc_type)

            await asyncio.sleep(config.polite_delay_seconds)

        return result


async def run_pipeline(config: Config) -> list[DownloadResult]:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    funds = load_funds(config.excel_file)
    logger.info("Loaded %d fund(s) from %s", len(funds), config.excel_file)

    semaphore = asyncio.Semaphore(config.max_concurrency)
    async with httpx.AsyncClient(cookies=config.cookies, verify=False) as client:
        results = await asyncio.gather(*[
            process_fund(fund, client, config, semaphore) for fund in funds
        ])

    write_results_to_excel(config.excel_file, list(results))
    return list(results)


def print_summary(results: list[DownloadResult]) -> None:
    print(f"\n{'='*60}")
    print("Summary:")
    for r in results:
        status = "OK" if r.downloaded_files else "FAILED"
        print(f"  [{status}] {r.fund.isin}: downloaded={sorted(r.downloaded_files.keys())} errors={r.errors}")


async def main() -> None:
    config = Config()
    try:
        results = await run_pipeline(config)
        print_summary(results)
    except Exception:
        logger.exception("Pipeline failed")
        raise


if __name__ == "__main__":
    asyncio.run(main())
