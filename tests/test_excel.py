import openpyxl
import pytest

from fund_mind import DownloadResult, Fund, load_funds, write_results_to_excel


def make_xlsx(tmp_path, headers, rows):
    path = tmp_path / "funds.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


class TestLoadFunds:
    def test_loads_isins(self, tmp_path):
        path = make_xlsx(tmp_path, ["IDINSTRUMENT", "DESCRIPTION"],
                         [["CH0002788708", "UBS Asia"], ["IE00BF4RFH31", "Some Fund"]])
        funds = load_funds(path)
        assert len(funds) == 2
        assert funds[0].isin == "CH0002788708"
        assert funds[0].name == "UBS Asia"
        assert funds[1].isin == "IE00BF4RFH31"

    def test_skips_blank_isin_rows(self, tmp_path):
        path = make_xlsx(tmp_path, ["IDINSTRUMENT"], [["CH0002788708"], [None], [""]])
        funds = load_funds(path)
        assert len(funds) == 1

    def test_missing_idinstrument_raises(self, tmp_path):
        path = make_xlsx(tmp_path, ["ISIN"], [["CH0002788708"]])
        with pytest.raises(ValueError, match="IDINSTRUMENT"):
            load_funds(path)

    def test_missing_description_defaults_to_empty(self, tmp_path):
        path = make_xlsx(tmp_path, ["IDINSTRUMENT"], [["CH0002788708"]])
        funds = load_funds(path)
        assert funds[0].name == ""

    def test_file_not_found_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            load_funds(tmp_path / "nonexistent.xlsx")


class TestWriteResultsToExcel:
    def _make_result(self, isin="CH0002788708"):
        return DownloadResult(
            fund=Fund(isin=isin, name="Test Fund"),
            downloaded_files={"PR": f"{isin}_PR_EN_2025-01-01.pdf", "AR": f"{isin}_AR_EN_2025-01-01.pdf"},
            legal_form="SICAV",
            ter_excl_performance_fee="0.0182",
            ept_valuation_frequency="252",
            ept_investment_objective="Growth",
            manco="Test ManCo",
            fund_domicile="LU",
            fund_launch_date="1997-03-14",
            fund_currency="USD",
        )

    def test_metadata_columns_written(self, tmp_path):
        path = make_xlsx(tmp_path, ["IDINSTRUMENT"], [["CH0002788708"]])
        write_results_to_excel(path, [self._make_result()])
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        row = {h: v for h, v in zip(headers, [c.value for c in ws[2]])}
        assert row["Legal Form"] == "SICAV"
        assert row["TER Excluding Performance Fee"] == "0.0182"
        assert row["ManCo"] == "Test ManCo"
        assert row["Fund Domicile Alpha-2"] == "LU"
        assert row["Fund Launch Date"] == "1997-03-14"
        assert row["Fund Currency"] == "USD"

    def test_doc_columns_written(self, tmp_path):
        path = make_xlsx(tmp_path, ["IDINSTRUMENT"], [["CH0002788708"]])
        write_results_to_excel(path, [self._make_result()])
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        row = {h: v for h, v in zip(headers, [c.value for c in ws[2]])}
        assert row["Prospectus"] == "CH0002788708_PR_EN_2025-01-01.pdf"
        assert row["Annual Report"] == "CH0002788708_AR_EN_2025-01-01.pdf"
        assert row["Monthly Report"] is None or row["Monthly Report"] == ""

    def test_rerunning_does_not_duplicate_columns(self, tmp_path):
        path = make_xlsx(tmp_path, ["IDINSTRUMENT"], [["CH0002788708"]])
        result = self._make_result()
        write_results_to_excel(path, [result])
        write_results_to_excel(path, [result])
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers.count("Legal Form") == 1
        assert headers.count("Prospectus") == 1

    def test_unknown_isin_row_left_blank(self, tmp_path):
        path = make_xlsx(tmp_path, ["IDINSTRUMENT"], [["CH0002788708"], ["XX0000000000"]])
        write_results_to_excel(path, [self._make_result("CH0002788708")])
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        row3 = {h: v for h, v in zip(headers, [c.value for c in ws[3]])}
        assert row3.get("Legal Form") is None or row3.get("Legal Form") == ""
